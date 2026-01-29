import argparse
import datetime as dt
import hashlib
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request

from openpyxl import load_workbook


CACHE_DIR = os.path.join(os.getcwd(), ".cache_market_data")
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"


def ensure_cache_dir():
    os.makedirs(CACHE_DIR, exist_ok=True)


def sha1(text):
    return hashlib.sha1(text.encode("utf-8")).hexdigest()


def read_cache(key, ttl_seconds=6 * 3600):
    ensure_cache_dir()
    path = os.path.join(CACHE_DIR, f"{sha1(key)}.json")
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        if time.time() - payload.get("ts", 0) > ttl_seconds:
            return None
        return payload.get("data")
    except Exception:
        return None


def write_cache(key, data):
    ensure_cache_dir()
    path = os.path.join(CACHE_DIR, f"{sha1(key)}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"ts": time.time(), "data": data}, f)


def http_get_json(url, timeout=20):
    cache_key = f"GET:{url}"
    cached = read_cache(cache_key)
    if cached is not None:
        return cached
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT, "Accept": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            text = resp.read().decode("utf-8", errors="ignore")
    except urllib.error.HTTPError as exc:
        text = exc.read().decode("utf-8", errors="ignore") if hasattr(exc, "read") else ""
        return {"_error": True, "status": exc.code, "body": text[:300], "url": url}
    except Exception as exc:
        return {"_error": True, "status": 0, "body": str(exc)[:300], "url": url}
    try:
        data = json.loads(text)
    except Exception:
        data = {"_error": True, "status": 0, "body": text[:300], "url": url}
    write_cache(cache_key, data)
    return data


def normalize_symbol(ticker):
    if not ticker:
        return ""
    raw = str(ticker).strip().upper()
    if "." in raw:
        return raw
    if re.match(r"^[A-Z]{4,6}\d{1,2}[A-Z]?$", raw):
        return f"{raw}.SA"
    return raw


def to_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(raw, fmt).date()
        except Exception:
            continue
    try:
        return dt.datetime.fromisoformat(raw).date()
    except Exception:
        return None


def to_float(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace(",", ".")
    try:
        return float(raw)
    except Exception:
        return None


def get_history(ticker, start_date, end_date):
    symbol = normalize_symbol(ticker)
    if not symbol or not start_date or not end_date:
        return None
    start_ts = int(dt.datetime.combine(start_date, dt.time.min).timestamp())
    end_ts = int(dt.datetime.combine(end_date, dt.time.max).timestamp())
    url = (
        "https://query1.finance.yahoo.com/v8/finance/chart/"
        f"{urllib.parse.quote(symbol)}?period1={start_ts}&period2={end_ts}&interval=1d&events=div"
    )
    data = http_get_json(url)
    if data.get("_error"):
        return {"error": data}
    result = (data.get("chart") or {}).get("result") or []
    if not result:
        return {"error": {"status": 502, "body": "empty result"}}
    return result[0]


def get_spot(ticker):
    symbol = normalize_symbol(ticker)
    if not symbol:
        return None, "none"
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}?interval=1d&range=5d"
    data = http_get_json(url)
    if data.get("_error"):
        return None, "yahoo"
    result = (data.get("chart") or {}).get("result") or []
    if not result:
        return None, "yahoo"
    meta = result[0].get("meta") or {}
    closes = ((result[0].get("indicators") or {}).get("quote") or [{}])[0].get("close") or []
    last_close = next((v for v in reversed(closes) if v is not None), None)
    spot = meta.get("regularMarketPrice", last_close)
    return spot, "yahoo"


def get_dividends(history, start_date, end_date):
    if not history:
        return 0.0
    events = history.get("events") or {}
    dividends = events.get("dividends") or {}
    total = 0.0
    for ts, item in dividends.items():
        try:
            ts_int = int(ts)
        except Exception:
            continue
        date = dt.datetime.utcfromtimestamp(ts_int).date()
        if start_date <= date <= end_date:
            amount = to_float(item.get("amount"))
            if amount:
                total += amount
    return total


def slice_history(history, start_date, end_date):
    if not history:
        return [], [], [], []
    timestamps = history.get("timestamp") or []
    quote = ((history.get("indicators") or {}).get("quote") or [{}])[0]
    highs = quote.get("high") or []
    lows = quote.get("low") or []
    closes = quote.get("close") or []
    highs_out, lows_out, closes_out = [], [], []
    for idx, ts in enumerate(timestamps):
        try:
            date = dt.datetime.utcfromtimestamp(int(ts)).date()
        except Exception:
            continue
        if start_date <= date <= end_date:
            high_val = highs[idx] if idx < len(highs) else None
            low_val = lows[idx] if idx < len(lows) else None
            close_val = closes[idx] if idx < len(closes) else None
            if high_val is not None:
                highs_out.append(high_val)
            if low_val is not None:
                lows_out.append(low_val)
            if close_val is not None:
                closes_out.append(close_val)
    return highs_out, lows_out, closes_out, timestamps


def detect_barrier_type(barrier_type):
    if not barrier_type:
        return None, None
    text = str(barrier_type).upper()
    mode = None
    if "OUT" in text or "KO" in text:
        mode = "out"
    if "IN" in text or "KI" in text:
        mode = "in" if mode is None else mode
    direction = None
    if "UP" in text or "UO" in text or "UI" in text:
        direction = "high"
    if "DOWN" in text or "DO" in text or "DI" in text:
        direction = "low"
    return mode, direction


def barrier_active(barrier_value, mode, direction, spot_initial, high_max, low_min):
    if barrier_value is None or mode is None:
        return None
    if direction is None and spot_initial is not None:
        direction = "high" if barrier_value >= spot_initial else "low"
    if direction == "high":
        activated = high_max is not None and high_max >= barrier_value
    else:
        activated = low_min is not None and low_min <= barrier_value
    return activated


def mask_id(value):
    if value is None:
        return "row"
    text = str(value)
    if len(text) <= 6:
        return f"{text[:2]}***"
    return f"{text[:3]}***{text[-3:]}"


def find_header_row(ws, max_rows=10):
    for r in range(1, max_rows + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 80)]
        if any(v is not None and str(v).strip() != "" for v in row_vals):
            return r
    return 1


def normalize_header(text):
    if text is None:
        return ""
    value = str(text).strip().lower()
    value = re.sub(r"[^a-z0-9()]+", "", value)
    return value


def build_header_map(ws, header_row):
    headers = {}
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c).value
        if cell is None or str(cell).strip() == "":
            continue
        headers[normalize_header(cell)] = c
    return headers


def get_col(headers, *keys):
    for key in keys:
        norm = normalize_header(key)
        if norm in headers:
            return headers[norm]
    return None


def get_leg_cols(headers, leg_index):
    suffix = f"({leg_index})"
    return {
        "qty": get_col(headers, f"Quantidade Ativa {suffix}", f"QuantidadeAtiva{suffix}"),
        "tipo": get_col(headers, f"Tipo {suffix}", f"Tipo{suffix}"),
        "strike": get_col(headers, f"Valor do Strike {suffix}", f"ValordoStrike{suffix}"),
        "barreira_valor": get_col(headers, f"Valor da Barreira {suffix}", f"ValordaBarreira{suffix}"),
        "barreira_tipo": get_col(headers, f"Tipo da Barreira {suffix}", f"TipodaBarreira{suffix}"),
        "rebate": get_col(headers, f"Valor do Rebate {suffix}", f"ValordoRebate{suffix}"),
        "multiplicador": get_col(headers, f"Multiplicador {suffix}", f"Multiplicador{suffix}"),
    }


def auto_find_input():
    candidates = []
    for root, _, files in os.walk(os.getcwd()):
        for name in files:
            if name.lower().endswith((".xlsx", ".xlsm")) and "vencimento" in name.lower():
                candidates.append(os.path.join(root, name))
    if candidates:
        return max(candidates, key=lambda p: os.path.getmtime(p))
    known = [
        r"C:\Users\endri\3D Objects\mesarv\Vencimentos\Relatório de Posição.xlsx",
        r"C:\Users\endri\PWR-Endrio\Vencimentos\Relatório de Posição.xlsx",
    ]
    existing = [p for p in known if os.path.exists(p)]
    if existing:
        return max(existing, key=lambda p: os.path.getmtime(p))
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", help="Caminho do arquivo .xlsx/.xlsm")
    parser.add_argument("--sheet", help="Nome da aba")
    parser.add_argument("--date", help="Data base YYYY-MM-DD")
    args = parser.parse_args()

    input_path = args.input or auto_find_input()
    if not input_path or not os.path.exists(input_path):
        print("Arquivo de entrada nao encontrado. Use --input.")
        sys.exit(1)

    base_date = dt.date.today()
    if args.date:
        base_date = dt.datetime.fromisoformat(args.date).date()

    keep_vba = input_path.lower().endswith(".xlsm")
    wb = load_workbook(input_path, data_only=True, keep_vba=keep_vba)
    sheet_name = args.sheet or wb.sheetnames[0]
    ws = wb[sheet_name]

    header_row = find_header_row(ws)
    headers = build_header_map(ws, header_row)

    col_ativo = get_col(headers, "Ativo")
    col_registro = get_col(headers, "Data Registro")
    col_venc = get_col(headers, "Data Vencimento")
    col_valor_ativo = get_col(headers, "Valor Ativo")
    col_custo_unitario = get_col(headers, "Custo Unitário Cliente")

    if not all([col_ativo, col_registro, col_venc]):
        print("Colunas obrigatorias nao encontradas (Ativo/Data Registro/Data Vencimento).")
        sys.exit(1)

    legs = {i: get_leg_cols(headers, i) for i in range(1, 5)}

    # create output columns
    new_headers = [
        "Entrada_Total",
        "Atual_Total",
        "PnL",
        "PnL_%",
        "Dividendos",
        "DTE",
        "Status_Vencimento",
        "Spot_Atual",
        "Fonte_Spot",
    ]
    start_col = ws.max_column + 1
    for idx, name in enumerate(new_headers):
        ws.cell(row=header_row, column=start_col + idx, value=name)

    log_rows = []
    rows_data = []
    ticker_ranges = {}

    for row in range(header_row + 1, ws.max_row + 1):
        ticker = ws.cell(row=row, column=col_ativo).value
        if ticker is None or str(ticker).strip() == "":
            continue

        data_registro = to_date(ws.cell(row=row, column=col_registro).value)
        data_venc = to_date(ws.cell(row=row, column=col_venc).value)
        if not data_registro or not data_venc:
            continue

        spot_entry = to_float(ws.cell(row=row, column=col_valor_ativo).value)
        end_date = min(base_date, data_venc)
        ticker_key = str(ticker).strip().upper()

        rows_data.append({
            "row": row,
            "ticker": ticker_key,
            "data_registro": data_registro,
            "data_venc": data_venc,
            "spot_entry": spot_entry,
        })

        if ticker_key not in ticker_ranges:
            ticker_ranges[ticker_key] = {"start": data_registro, "end": end_date}
        else:
            ticker_ranges[ticker_key]["start"] = min(ticker_ranges[ticker_key]["start"], data_registro)
            ticker_ranges[ticker_key]["end"] = max(ticker_ranges[ticker_key]["end"], end_date)

    history_map = {}
    spot_map = {}
    for ticker_key, window in ticker_ranges.items():
        history_map[ticker_key] = get_history(ticker_key, window["start"], window["end"])
        spot_map[ticker_key] = get_spot(ticker_key)

    for item in rows_data:
        row = item["row"]
        ticker = item["ticker"]
        data_registro = item["data_registro"]
        data_venc = item["data_venc"]
        spot_entry = item["spot_entry"]
        custo_unitario = to_float(ws.cell(row=row, column=col_custo_unitario).value) if col_custo_unitario else None

        end_date = min(base_date, data_venc)
        history = history_map.get(ticker)
        history_error = history.get("error") if isinstance(history, dict) else None

        highs_clean, lows_clean, closes_clean, _ = slice_history(history, data_registro, end_date)
        high_max = max(highs_clean) if highs_clean else None
        low_min = min(lows_clean) if lows_clean else None
        last_close = closes_clean[-1] if closes_clean else None

        spot_current, spot_source = spot_map.get(ticker, (None, "yahoo"))
        spot_for_payoff = last_close if data_venc < base_date else spot_current

        entrada_total = 0.0
        atual_total = 0.0
        dividendos_total = 0.0
        barrier_status = []

        for i in range(1, 5):
            cols = legs[i]
            qty = to_float(ws.cell(row=row, column=cols["qty"]).value) if cols["qty"] else None
            tipo = ws.cell(row=row, column=cols["tipo"]).value if cols["tipo"] else None
            strike = to_float(ws.cell(row=row, column=cols["strike"]).value) if cols["strike"] else None
            barreira_valor = to_float(ws.cell(row=row, column=cols["barreira_valor"]).value) if cols["barreira_valor"] else None
            barreira_tipo = ws.cell(row=row, column=cols["barreira_tipo"]).value if cols["barreira_tipo"] else None
            rebate = to_float(ws.cell(row=row, column=cols["rebate"]).value) if cols["rebate"] else None
            multiplicador = to_float(ws.cell(row=row, column=cols["multiplicador"]).value) if cols["multiplicador"] else None

            if qty in (None, 0):
                continue
            tipo_text = str(tipo).upper() if tipo is not None else ""

            is_stock = "ESTOQUE" in tipo_text or "ACAO" in tipo_text or "AÇÃO" in tipo_text or "STOCK" in tipo_text
            is_call = "CALL" in tipo_text
            is_put = "PUT" in tipo_text

            if is_stock:
                unit_entry = spot_entry or 0.0
                entrada_total += qty * unit_entry
                if spot_current is not None:
                    atual_total += qty * spot_current
                dividendos_total += qty * get_dividends(history, data_registro, end_date)
                continue

            if not (is_call or is_put):
                continue

            mult = multiplicador if multiplicador not in (None, 0) else 100.0
            entry_unit = abs(custo_unitario) if custo_unitario is not None else 0.0
            entrada_total += qty * entry_unit

            payoff_unit = None
            if spot_for_payoff is not None and strike is not None:
                if is_call:
                    payoff_unit = max(0.0, spot_for_payoff - strike)
                if is_put:
                    payoff_unit = max(0.0, strike - spot_for_payoff)

            mode, direction = detect_barrier_type(barreira_tipo)
            activated = barrier_active(barreira_valor, mode, direction, spot_entry, high_max, low_min)
            if mode == "out" and activated:
                payoff_unit = 0.0
                if rebate not in (None, 0):
                    payoff_unit = rebate
                barrier_status.append(f"OUT_{i}")
            elif mode == "in" and activated is False:
                payoff_unit = 0.0
                barrier_status.append(f"IN_NAO_{i}")
            else:
                if mode in ("in", "out"):
                    barrier_status.append(f"{mode.upper()}_{i}")

            if payoff_unit is not None:
                atual_total += qty * mult * payoff_unit

        pnl = atual_total - entrada_total
        pnl_pct = (pnl / entrada_total) if entrada_total else None
        dte = (data_venc - base_date).days
        status_venc = "VENCIDO" if dte < 0 else "ATIVO"

        write_values = [
            entrada_total,
            atual_total,
            pnl,
            pnl_pct,
            dividendos_total,
            dte,
            status_venc,
            spot_for_payoff,
            spot_source,
        ]
        for idx, value in enumerate(write_values):
            ws.cell(row=row, column=start_col + idx, value=value)

        log_rows.append({
            "id": mask_id(ws.cell(row=row, column=3).value),
            "ticker": str(ticker).strip().upper(),
            "entrada_total": round(entrada_total, 2),
            "atual_total": round(atual_total, 2),
            "pnl": round(pnl, 2),
            "dte": dte,
            "status_venc": status_venc,
            "barreiras": barrier_status,
            "spot": spot_for_payoff if spot_for_payoff is None else round(spot_for_payoff, 4),
            "source": spot_source,
            "history_error": history_error,
        })

    base_name, ext = os.path.splitext(os.path.basename(input_path))
    date_suffix = base_date.strftime("%Y-%m-%d")
    output_name = f"{base_name}_atualizada_{date_suffix}{ext if ext else '.xlsx'}"
    output_path = os.path.join(os.path.dirname(input_path), output_name)
    wb.save(output_path)

    log_name = f"log_resumo_{date_suffix}.json"
    log_path = os.path.join(os.path.dirname(input_path), log_name)
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump({"rows": log_rows, "total": len(log_rows)}, f, ensure_ascii=False, indent=2)

    print(f"Arquivo atualizado: {output_path}")
    print(f"Log: {log_path}")
    print(f"Linhas processadas: {len(log_rows)}")


if __name__ == "__main__":
    main()
